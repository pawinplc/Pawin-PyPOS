from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date, timedelta
from io import BytesIO
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_db
from app.models.models import User, Sale, SaleItem, Item, StockMovement
from app.routers.auth import get_current_active_user

router = APIRouter(prefix="/api/reports", tags=["Reports"])

def create_pdf_buffer(title, headers, data, column_widths):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1
    )
    
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))
    
    table = Table([headers] + data, colWidths=column_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def create_excel_buffer(title, headers, data):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        ws.column_dimensions[column].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@router.get("/daily-sales/pdf")
def daily_sales_pdf(
    date: str = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if date:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
    else:
        target_date = date.today()
    
    start_datetime = datetime.combine(target_date, datetime.min.time())
    end_datetime = datetime.combine(target_date, datetime.max.time())
    
    sales = db.query(Sale).filter(
        Sale.created_at >= start_datetime,
        Sale.created_at <= end_datetime
    ).all()
    
    headers = ["Receipt No", "Time", "Cashier", "Items", "Total", "Discount", "Final"]
    data = []
    grand_total = Decimal("0.00")
    
    for sale in sales:
        sale_items = db.query(SaleItem).filter(SaleItem.sale_id == sale.id).all()
        items_count = sum(si.quantity for si in sale_items)
        grand_total += sale.final_amount
        data.append([
            f"#{sale.id:05d}",
            sale.created_at.strftime("%H:%M"),
            sale.cashier.full_name or sale.cashier.username,
            items_count,
            f"฿{sale.total_amount:,.2f}",
            f"฿{sale.discount_amount:,.2f}",
            f"฿{sale.final_amount:,.2f}"
        ])
    
    data.append(["", "", "", "GRAND TOTAL:", "", f"฿{grand_total:,.2f}"])
    
    title = f"Daily Sales Report - {target_date.strftime('%d %B %Y')}"
    buffer = create_pdf_buffer(
        title,
        headers,
        data,
        [80, 60, 100, 60, 80, 80, 80]
    )
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=daily_sales_{date or 'today'}.pdf"}
    )

@router.get("/daily-sales/excel")
def daily_sales_excel(
    date: str = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if date:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
    else:
        target_date = date.today()
    
    start_datetime = datetime.combine(target_date, datetime.min.time())
    end_datetime = datetime.combine(target_date, datetime.max.time())
    
    sales = db.query(Sale).filter(
        Sale.created_at >= start_datetime,
        Sale.created_at <= end_datetime
    ).all()
    
    headers = ["Receipt No", "Date", "Time", "Cashier", "Items", "Total", "Discount", "Final"]
    data = []
    grand_total = Decimal("0.00")
    
    for sale in sales:
        sale_items = db.query(SaleItem).filter(SaleItem.sale_id == sale.id).all()
        items_count = sum(si.quantity for si in sale_items)
        grand_total += sale.final_amount
        data.append([
            f"#{sale.id:05d}",
            sale.created_at.strftime("%Y-%m-%d"),
            sale.created_at.strftime("%H:%M"),
            sale.cashier.full_name or sale.cashier.username,
            items_count,
            float(sale.total_amount),
            float(sale.discount_amount),
            float(sale.final_amount)
        ])
    
    data.append(["", "", "", "GRAND TOTAL:", "", "", "", float(grand_total)])
    
    title = f"Daily Sales {target_date.strftime('%Y%m%d')}"
    buffer = create_excel_buffer(title, headers, data)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=daily_sales_{date or 'today'}.xlsx"}
    )

@router.get("/monthly-sales/pdf")
def monthly_sales_pdf(
    year: int = Query(None),
    month: int = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not year:
        year = date.today().year
    if not month:
        month = date.today().month
    
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    sales = db.query(Sale).filter(
        Sale.created_at >= start_date,
        Sale.created_at < end_date
    ).all()
    
    headers = ["Date", "Receipts", "Total Sales", "Final Amount"]
    data = []
    daily_totals = {}
    
    for sale in sales:
        day_key = sale.created_at.strftime("%Y-%m-%d")
        if day_key not in daily_totals:
            daily_totals[day_key] = {"count": 0, "total": Decimal("0.00"), "final": Decimal("0.00")}
        daily_totals[day_key]["count"] += 1
        daily_totals[day_key]["total"] += sale.total_amount
        daily_totals[day_key]["final"] += sale.final_amount
    
    grand_total = Decimal("0.00")
    for day_key in sorted(daily_totals.keys()):
        stats = daily_totals[day_key]
        grand_total += stats["final"]
        data.append([
            day_key,
            stats["count"],
            f"฿{stats['total']:,.2f}",
            f"฿{stats['final']:,.2f}"
        ])
    
    data.append(["GRAND TOTAL", len(sales), "", f"฿{grand_total:,.2f}"])
    
    month_name = date(year, month, 1).strftime("%B %Y")
    title = f"Monthly Sales Report - {month_name}"
    buffer = create_pdf_buffer(title, headers, data, [100, 80, 120, 120])
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=monthly_sales_{year}_{month:02d}.pdf"}
    )

@router.get("/monthly-sales/excel")
def monthly_sales_excel(
    year: int = Query(None),
    month: int = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not year:
        year = date.today().year
    if not month:
        month = date.today().month
    
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    sales = db.query(Sale).filter(
        Sale.created_at >= start_date,
        Sale.created_at < end_date
    ).all()
    
    headers = ["Date", "Receipts", "Total Sales", "Discount", "Final Amount"]
    data = []
    daily_totals = {}
    
    for sale in sales:
        day_key = sale.created_at.strftime("%Y-%m-%d")
        if day_key not in daily_totals:
            daily_totals[day_key] = {"count": 0, "total": Decimal("0.00"), "discount": Decimal("0.00"), "final": Decimal("0.00")}
        daily_totals[day_key]["count"] += 1
        daily_totals[day_key]["total"] += sale.total_amount
        daily_totals[day_key]["discount"] += sale.discount_amount
        daily_totals[day_key]["final"] += sale.final_amount
    
    grand_total = Decimal("0.00")
    for day_key in sorted(daily_totals.keys()):
        stats = daily_totals[day_key]
        grand_total += stats["final"]
        data.append([
            day_key,
            stats["count"],
            float(stats["total"]),
            float(stats["discount"]),
            float(stats["final"])
        ])
    
    data.append(["GRAND TOTAL", len(sales), "", "", float(grand_total)])
    
    title = f"Monthly Sales {year}_{month:02d}"
    buffer = create_excel_buffer(title, headers, data)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=monthly_sales_{year}_{month:02d}.xlsx"}
    )

@router.get("/stock-arrival/pdf")
def stock_arrival_pdf(
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view stock arrival reports")
    
    query = db.query(StockMovement).filter(StockMovement.movement_type == "in")
    
    if start_date:
        start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
        query = query.filter(StockMovement.created_at >= start_datetime)
    if end_date:
        end_datetime = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query = query.filter(StockMovement.created_at < end_datetime)
    
    movements = query.order_by(StockMovement.created_at.desc()).all()
    
    headers = ["Date", "Item", "SKU", "Quantity", "Reference", "Notes", "By"]
    data = []
    
    for m in movements:
        data.append([
            m.created_at.strftime("%Y-%m-%d"),
            m.item.name if m.item else "N/A",
            m.item.sku if m.item else "N/A",
            m.quantity,
            m.reference or "-",
            m.notes or "-",
            m.user.username if m.user else "N/A"
        ])
    
    title = "Stock Arrival Report"
    if start_date and end_date:
        title += f" ({start_date} to {end_date})"
    elif start_date:
        title += f" (from {start_date})"
    elif end_date:
        title += f" (until {end_date})"
    
    buffer = create_pdf_buffer(title, headers, data, [80, 150, 80, 60, 100, 150, 80])
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=stock_arrival.pdf"}
    )

@router.get("/stock-arrival/excel")
def stock_arrival_excel(
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view stock arrival reports")
    
    query = db.query(StockMovement).filter(StockMovement.movement_type == "in")
    
    if start_date:
        start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
        query = query.filter(StockMovement.created_at >= start_datetime)
    if end_date:
        end_datetime = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        query = query.filter(StockMovement.created_at < end_datetime)
    
    movements = query.order_by(StockMovement.created_at.desc()).all()
    
    headers = ["Date", "Time", "Item", "SKU", "Category", "Quantity", "Unit Cost", "Total Cost", "Reference", "Notes", "Received By"]
    data = []
    
    for m in movements:
        total_cost = m.quantity * (m.item.cost_price if m.item else 0)
        data.append([
            m.created_at.strftime("%Y-%m-%d"),
            m.created_at.strftime("%H:%M"),
            m.item.name if m.item else "N/A",
            m.item.sku if m.item else "N/A",
            m.item.category.name if m.item and m.item.category else "N/A",
            m.quantity,
            float(m.item.cost_price) if m.item else 0,
            float(total_cost),
            m.reference or "-",
            m.notes or "-",
            m.user.username if m.user else "N/A"
        ])
    
    title = "Stock Arrival Report"
    buffer = create_excel_buffer(title, headers, data)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stock_arrival.xlsx"}
    )
